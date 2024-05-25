from asyncio.windows_events import NULL
import encodings
import json
from msilib.schema import File
from pickle import FRAME
from queue import Full
from tkinter import YES
from venv import create
from fpdf.util import Padding
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import rows_from_range
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Fill, PatternFill, NamedStyle, borders
from fpdf import FPDF
import smtplib
import os
import mimetypes
from email import encoders
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
from email.mime.audio import MIMEAudio
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from dotenv import load_dotenv
import shutil
import re

load_dotenv()

def create_pdf(change, send_to):
    
    if change == 2:
        name_doc = 'БЛАГОДАРСТВЕННОЕ\nПИСЬМО'
    else:
        name_doc = 'СЕРТИФИКАТ'
    
    workBook = load_workbook(filename = 'Таблица/Таблица.xlsx')
    sheets = workBook.active
    
    full_name = ''
    code = ''
    mail = ''

    for row in range(1, sheets.max_row + 1):
        
        if send_to.lower() == 'да':
            os.mkdir('To_Send')
        
        pdf = FPDF('P', 'mm', 'A4')
        
        pdf.add_font("Sans", style="", fname="Noto Sans/NotoSans-Regular.ttf", uni=True)
        pdf.add_font("Sans", style="B", fname="Noto Sans/NotoSans-Bold.ttf", uni=True)
        pdf.add_font("Sans", style="I", fname="Noto Sans/NotoSans-Italic.ttf", uni=True)
        pdf.add_font("Sans", style="BI", fname="Noto Sans/NotoSans-BoldItalic.ttf", uni=True)
        
        fist_name = sheets[row][0].value
        if fist_name == None:
            fist_name = ''
        
        name = sheets[row][1].value
        if name == None:
            name = ''
        
        second_name = sheets[row][2].value
        if second_name == None:
            second_name = ''
        
        code = sheets[row][3].value
        if code == None:
            code = ''
        
        mail = sheets[row][4].value
        if mail == None:
            mail = ''
    
        full_name = fist_name + ' ' + name + ' ' + second_name
    
        print(row, full_name, code, mail)
        
        if change == 1:
            pdf.set_page_background('Изображение/СЕРТИФИКАТ.png')
        else:
            pdf.set_page_background('Изображение/Благодарственное письмо.png')
            
        pdf.add_page()
        
        pdf.set_font("Sans", 'B', 35)
        pdf.set_text_color(0, 40, 100)
        pdf.y = 145
        pdf.multi_cell(0, 15, text = name_doc, border = 0, align = 'C')
        pdf.ln()
        
        pdf.y = 180
        pdf.set_font("Sans", 'B', 24)
        pdf.set_text_color(70, 70, 70)
        
        if change == 2:
            pdf.multi_cell(0, 10, text = 'Уважаемый\n' + full_name, border = 0, align = 'C')
            pdf.ln()
        else:
            pdf.multi_cell(0, 10, text = 'Получил\n' + full_name, border = 0, align = 'C')
            pdf.ln()
        
        pdf.set_font("Sans", '', 14)
        pdf.y = 205
        pdf.x = 10
        pdf.multi_cell(190, h = 6, text = code, border = 0, align='C')
        pdf.ln()
        
        text_file = open('Таблица/TextFile.txt', 'r')
        get_text = text_file.read()
        text_file.close()
        
        pdf.set_font("Sans", '', 14)
        pdf.y = 225
        pdf.x = 15
        pdf.multi_cell(180, h = 6, text = get_text, border = 0, align='C')
        pdf.ln()
        
        pdf.set_font("Sans", 'B', 16)
        pdf.y = 260
        pdf.cell(0, 6, text = 'Ректор НГПУ ______________________ А.А. Галиакберова', border = 0, align='C')
        pdf.ln()

        pdf.set_display_mode(zoom='fullpage', layout='continuous')
        
        if send_to.lower() == 'да':
            pdf.output("To_Send/" + full_name + ".pdf")
            
            sender = "msboriskv@gmail.com"
            password = os.getenv("point")
            
            msg = MIMEMultipart()
            msg["From"] = sender
            msg["To"] = mail
            msg["Subject"] = 'Сертификат форума "Образование: Реалии и перспективы"'

            msg.attach(MIMEText("Здравствуйте, " + full_name + " отправляем Вам сертификат форума - Образование: Реалии и перспективы"))

            for file in os.listdir("To_Send"):
                filename = os.path.basename(file)
                ftype, encoding = mimetypes.guess_type(file)
                file_type, subtype = ftype.split("/")
    
                if file_type == "text":
                    with open(f"To_Send/{file}") as f:
                        file = MIMEText(f.read())
                elif file_type == "image":
                    with open(f"To_Send/{file}", "rb") as f:
                        file = MIMEImage(f.read(), subtype)
                elif file_type == "audio":
                    with open(f"To_Send/{file}", "rb") as f:
                        file = MIMEAudio(f.read(), subtype)
                elif file_type == "application":
                    with open(f"To_Send/{file}", "rb") as f:
                        file = MIMEApplication(f.read(), subtype)
                else:
                    with open(f"To_Send"/{file}, "rb") as f:
                        file = MIMEBase(file_type, subtype)
                        file.set_payload(f.read())
                        encoders.encode_base64(file)
    
                file.add_header('Content-Disposition', 'attachment', filename=filename)
                msg.attach(file)

            server = smtplib.SMTP("smtp.gmail.com", 587)
            server.starttls()

            server.login(sender, password)
            server.sendmail(sender, mail, msg.as_string())
            
            shutil.rmtree('To_Send')
            
            #pdf.output("PDF_Печать/" + full_name + ".pdf")
            
        else:
            pdf.output("PDF_Печать/" + full_name + ".pdf")

def main():

    change = int()
    send_to = ''

    while change == 1 or 2:
        change = int(input('Введите цифру 1 для получение сертификата, цифру 2 для благодарственного письма: '))
        send_to = input('Если необходимо отправить документ по электронной почте введите (ДА/НЕТ): ')
        if not 1 <= change <= 2:
            print("Принимаются только числа от 1 до 2, попробуйте еще раз")
        else:
            shutil.rmtree('PDF_Печать/')
            os.mkdir('PDF_Печать/')
    
            return(create_pdf(change, send_to))

    

if __name__ == "__main__":
    main()