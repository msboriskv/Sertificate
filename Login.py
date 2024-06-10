﻿import sys
from PyQt6.QtWidgets import QApplication, QCheckBox, QWidget, QLabel, QLineEdit, \
    QPushButton, QVBoxLayout, QMessageBox, QMainWindow, QTabWidget, \
    QFileDialog, QTextEdit
from PyQt6.QtGui import QPixmap
from PyQt6.QtCore import Qt
from fpdf import FPDF
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, \
    Alignment, Font, NamedStyle, borders
import smtplib
import os
import mimetypes
from email import encoders
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
from email.mime.audio import MIMEAudio
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from dotenv import load_dotenv
import shutil
import re
import datetime
from datetime import datetime


load_dotenv()

if getattr(sys, 'frozen', False):
    # Если приложение собрано с помощью PyInstaller
    application_path = os.path.dirname(sys.executable)
else:
    # Если приложение запускается в интерпретаторе Python
    application_path = os.path.dirname(os.path.abspath(__file__))

class LoginWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('Сертификатор')
        self.setGeometry(800, 400, 280, 150)

        layout = QVBoxLayout()

        self.label_username = QLabel('Логин:')
        layout.addWidget(self.label_username)

        self.textbox_username = QLineEdit(self)
        layout.addWidget(self.textbox_username)

        self.label_password = QLabel('Пароль:')
        layout.addWidget(self.label_password)

        self.textbox_password = QLineEdit(self)
        self.textbox_password.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addWidget(self.textbox_password)

        self.button_login = QPushButton('Подтвердить', self)
        self.button_login.clicked.connect(self.check_login)
        layout.addWidget(self.button_login)

        self.setLayout(layout)

    def check_login(self):
        username = self.textbox_username.text()
        password = self.textbox_password.text()

        # Пример проверки логина и пароля
        if username == 'nicngpu@gmail.com' and password == '8c7raut?':
            QMessageBox.information(self, 'Успешно', 'Логин подтвержден')
            self.new_window = NewWindow()
            self.new_window.show()
            self.close()
            
        else:
            QMessageBox.warning(self, 'Ошибка', 'Неправильный логин или пароль')
           

def is_valid_email(email):
    # Регулярное выражение для проверки корректности email
    regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
    
    if email is None or email == '':
        return False
    
    if re.match(regex, email):
        return True
    else:
        return False

invalid_chars = "\"/«»!@#$%^&*~`.,:;|?{}0123456789+-=_()"

def is_valid_string(s):
    # Проверка длины строки
    if not (2 <= len(s) <= 25):
        return False
    
    # Проверка на наличие недопустимых символов
    pattern = f"[{re.escape(invalid_chars)}]"
    
    if re.search(pattern, s):
        return False
    return True

def is_valid_profession(s):
    # Проверка длины строки
    if not (3 <= len(s) <= 250):
        return False
    return True
    
class NewWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Сертификатор")
        self.setGeometry(700, 200, 420, 800)

        # Создаем QTabWidget
        self.tab_widget = QTabWidget()
        
        # Добавляем вкладку для открытия файла
        self.image_tab = QWidget()
        self.tabl_tab = QWidget()
        self.text_tab = QWidget()
        self.mail_tab = QWidget()

        self.image_layout = QVBoxLayout()
        self.tabl_layout = QVBoxLayout()
        self.text_layout = QVBoxLayout()
        self.mail_layout = QVBoxLayout()
        
        #Добавляем вкладку с фоновым изображением бланка
        self.buttom_image = QPushButton("Выберите фон для сетификата")
        self.buttom_image.clicked.connect(self.open_image_file)
        self.image_layout.addWidget(self.buttom_image)
        self.image_tab.setLayout(self.image_layout)
        self.tab_widget.addTab(self.image_tab, "Изображение")
        
        #Добавляем вкладку с таблицей
        self.tabl_tab.setLayout(self.tabl_layout)
        self.tab_widget.addTab(self.tabl_tab, "Таблица")
        
        self.buttom_tabl = QPushButton("Выберите таблицу")
        self.buttom_tabl.clicked.connect(self.open_file_xls)
        self.tabl_layout.addWidget(self.buttom_tabl)
        
        self.file_path_image_edit = QLineEdit(self.image_tab)
        self.image_layout.addWidget(self.file_path_image_edit)
        
        self.file_path_edit = QLineEdit(self.tabl_tab)
        self.tabl_layout.addWidget(self.file_path_edit)
        
        self.instruction_tabl = QLabel("Твблица должна содержать следующую структуру:\n\
                                       \nПервая колонка - Фамилия\nВторая - Имя\
                                       \nТретья - Отчество\nЧетвертая - Место работы\
                                       \nПятая - Аддес электронной почты")
        self.tabl_layout.addWidget(self.instruction_tabl)

        self.setLayout(self.tabl_layout)
        
        # Добавляем вкладку с текстом
        self.text_tab.setLayout(self.text_layout)
        self.tab_widget.addTab(self.text_tab, "Текст")
        
        self.create_text_baze = QLabel("Введите название документа (Сертификат, Благодарственное письмо и т.д.)")
        self.text_layout.addWidget(self.create_text_baze)
        
        self.text_edit_baze = QTextEdit(self.text_tab)
        self.text_layout.addWidget(self.text_edit_baze)
        self.text_tab.setLayout(self.text_layout)
        
        self.create_dia_freand = QLabel("Введите обращение (Уважаемый, получил и т.д.)")
        self.text_layout.addWidget(self.create_dia_freand)
        
        self.text_dia_freand = QTextEdit(self.text_tab)
        self.text_layout.addWidget(self.text_dia_freand)
        self.text_tab.setLayout(self.text_layout)
        
        self.create_text_all = QLabel("Введите основной текст (кнопка Enter переносит текст на новую строку)")
        self.text_layout.addWidget(self.create_text_all)

        self.text_edit = QTextEdit(self.text_tab)
        self.text_layout.addWidget(self.text_edit)
        self.text_tab.setLayout(self.text_layout)
        
        # Создаем вкладку для заполнения электронного письма
        self.mail_tab.setLayout(self.mail_layout)
        self.tab_widget.addTab(self.mail_tab, "Почта")
        
        self.create_mail_subject = QLabel("Введите тему письма")
        self.mail_layout.addWidget(self.create_mail_subject)
        
        self.text_mail_subject = QTextEdit(self.mail_tab)
        self.mail_layout.addWidget(self.text_mail_subject)
        self.mail_tab.setLayout(self.mail_layout)
        
        self.create_mail_text = QLabel("Введите содержание письма")
        self.mail_layout.addWidget(self.create_mail_text)
        
        self.text_mail_text = QTextEdit(self.mail_tab)
        self.mail_layout.addWidget(self.text_mail_text)
        self.mail_tab.setLayout(self.mail_layout)

        # Создаем кнопку для создания сертификата
        self.create_button = QPushButton("Создать сертификат")
        self.create_button.clicked.connect(self.create_pdf)
        
        # Создаем галочку для отправки почты
        self.mail_button = QCheckBox("Отправить по почте")
        
        # Создаем кнопку для закрытия окна
        self.close_button = QPushButton("Выйти из приложения")
        self.close_button.clicked.connect(self.close)

        # Создаем основной макет и добавляем в него вкладки QTabWidget и кнопку
        self.main_layout = QVBoxLayout()
        self.main_layout.addWidget(self.tab_widget)
        self.main_layout.addWidget(self.mail_button)
        self.main_layout.addWidget(self.create_button)
        self.main_layout.addWidget(self.close_button)
        
        # Создаем контейнерный виджет и устанавливаем в него основной макет
        self.container = QWidget()
        self.container.setLayout(self.main_layout)

        # Устанавливаем контейнерный виджет в качестве центрального виджета окна
        self.setCentralWidget(self.container)
        
        # Добавляем QLabel для отображения изображения
        self.image_label = QLabel()
        self.image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)  # Центрирование изображения
        self.image_layout.addWidget(self.image_label)
    
    def open_image_file(self):
        self.filename_image_path, _ = QFileDialog.getOpenFileName(self, "Выберите изображение", \
                                                  "", "Images (*.png *.xpm *.jpg *.jpeg *.bmp *.gif)")
        if self.filename_image_path:
            self.file_path_image_edit.setText(self.filename_image_path)
            print(self.filename_image_path)
            self.pixmap = QPixmap(self.filename_image_path)
            self.resize_image()

    def open_file_xls(self):
        self.file_path_tabl, _ = QFileDialog.getOpenFileName(self, "Выберите файл", "", "Excel Files (*.xls *.xlsx)")
        if self.file_path_tabl:
            self.file_path_edit.setText(self.file_path_tabl)
            print(self.file_path_tabl)

    def resize_image(self):
        if hasattr(self, 'pixmap') and not self.pixmap.isNull():
            # Масштабируем изображение с сохранением пропорций
            scaled_pixmap = self.pixmap.scaled(self.image_label.size(), \
                                               Qt.AspectRatioMode.KeepAspectRatio, \
                                               Qt.TransformationMode.SmoothTransformation)
            self.image_label.setPixmap(scaled_pixmap)

    def resizeEvent(self, event):
        self.resize_image()
        super().resizeEvent(event)
        
    def showFileDialog(self):
        file_path, _ = QFileDialog.getOpenFileName(self, 'Open file', '', 'All Files (*)')
        if file_path:
            self.file_path_edit.setText(file_path)
            print(file_path)

    def create_pdf(self):
        timer = datetime.now()
                
        date_now = str(timer.day) + '.' +str(timer.month) + '.' + \
            str(timer.year) + '_' + str(timer.hour) + '.' + str(timer.minute) + '.' + str(timer.second)

        name_doc = self.text_edit_baze.toPlainText()
        
        current_directory = application_path
        directory_path = os.path.join(current_directory, "Готовые сертификаты_" + date_now)

        if not os.path.exists(directory_path):
            os.mkdir(directory_path)
        
        workBook = load_workbook(filename = self.file_path_tabl)
        sheets = workBook.active
    
        full_name = ''
        code = ''
        mail = ''
        error_emails = []


        for row in range(1, sheets.max_row + 1):
        
            to_send_directory = os.path.join(current_directory, "To_Send")
    
            if self.mail_button.isChecked() or not os.path.exists(to_send_directory):
                os.makedirs(to_send_directory, exist_ok=True)
            
            pdf = FPDF('P', 'mm', 'A4')
        
            pdf.add_font("Sans", style="", fname="Noto Sans/NotoSans-Regular.ttf", uni=True)
            pdf.add_font("Sans", style="B", fname="Noto Sans/NotoSans-Bold.ttf", uni=True)
            pdf.add_font("Sans", style="I", fname="Noto Sans/NotoSans-Italic.ttf", uni=True)
            pdf.add_font("Sans", style="BI", fname="Noto Sans/NotoSans-BoldItalic.ttf", uni=True)

            fist_name = sheets[row][0].value
            if fist_name == None:
                fist_name = ''
        
            name = sheets[row][1].value
            if name == None:
                name = ''
        
            second_name = sheets[row][2].value
            if second_name == None:
                second_name = ''
        
            code = sheets[row][3].value
            if code == None:
                code = ''
        
            mail = sheets[row][4].value
            if mail == None:
                mail = ''
    
            full_name = fist_name + ' ' + name + ' ' + second_name
            
            error_emails.clear()
            
            if not is_valid_email(mail) or not is_valid_string(fist_name) or not is_valid_string(name) \
                or not is_valid_string(second_name) or not is_valid_profession(code):
                error_emails.append(full_name + ": " + mail)
                with open(os.path.join(application_path, 'Список ошибок_' + date_now + '.txt'), 'a') as error_file:
                    for error in error_emails:
                        error_file.write(f"{error}\n")
                continue  # Пропустить этот адрес и перейти к следующему
            
            #print(row, full_name, code, mail)
            
            pdf.set_page_background(self.filename_image_path)
            
            pdf.add_page()
        
            pdf.set_font("Sans", 'B', 35)
            pdf.set_text_color(0, 40, 100)
            pdf.y = 145
            pdf.multi_cell(0, 15, text = name_doc, border = 0, align = 'C')
            pdf.ln()
        
            pdf.y = 180
            pdf.set_font("Sans", 'B', 24)
            pdf.set_text_color(70, 70, 70)
            pdf.multi_cell(0, 10, text = self.text_dia_freand.toPlainText() + "\n" + full_name, border = 0, align = 'C')
                
            pdf.set_font("Sans", 'B', 14)
            pdf.y = 205
            pdf.x = 10
            pdf.multi_cell(190, h = 6, text = code, border = 0, align='C')
            pdf.ln()
                
            pdf.set_font("Sans", '', 14)
            pdf.y = 225
            pdf.x = 15
            pdf.multi_cell(180, h = 6, text = self.text_edit.toPlainText(), border = 0, align='C')
            pdf.ln()
        
            pdf.set_font("Sans", 'B', 16)
            pdf.y = 260
            pdf.cell(0, 6, text = 'Ректор НГПУ ______________________ А.А. Галиакберова', border = 0, align='C')
            pdf.ln()

            pdf.set_display_mode(zoom ='fullpage', layout ='continuous')
        
            if self.mail_button.isChecked():
                pdf.output("To_Send/" + full_name + ".pdf")
            
                sender = "nicngpu@gmail.com"
                password = os.getenv("point")
            
                msg = MIMEMultipart()
                msg["From"] = sender
                msg["To"] = mail
                msg["Subject"] = self.text_mail_subject.toPlainText()

                msg.attach(MIMEText(self.text_mail_text.toPlainText()))

                for file in os.listdir("To_Send"):
                    filename = os.path.basename(file)
                    ftype, encoding = mimetypes.guess_type(file)
                    file_type, subtype = ftype.split("/")
    
                    if file_type == "text":
                        with open(f"To_Send/{file}") as f:
                            file = MIMEText(f.read())
                    elif file_type == "image":
                        with open(f"To_Send/{file}", "rb") as f:
                            file = MIMEImage(f.read(), subtype)
                    elif file_type == "audio":
                        with open(f"To_Send/{file}", "rb") as f:
                            file = MIMEAudio(f.read(), subtype)
                    elif file_type == "application":
                        with open(f"To_Send/{file}", "rb") as f:
                            file = MIMEApplication(f.read(), subtype)
                    else:
                        with open(f"To_Send"/{file}, "rb") as f:
                            file = MIMEBase(file_type, subtype)
                            file.set_payload(f.read())
                            encoders.encode_base64(file)
    
                    file.add_header('Content-Disposition', 'attachment', filename=filename)
                    msg.attach(file)

                server = smtplib.SMTP("smtp.gmail.com", 587)
                server.starttls()

                server.login(sender, password)
                server.sendmail(sender, mail, msg.as_string())
                
            else:
                pdf_file_path = os.path.join(directory_path, f"{full_name}.pdf")
                pdf.output(pdf_file_path)
            
            shutil.rmtree('To_Send')


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = LoginWindow()
    window.show()
    sys.exit(app.exec())