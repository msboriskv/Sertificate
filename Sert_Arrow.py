# -*- coding: cp1251 -*-
# -*- coding: utf8 -*-

from asyncio.windows_events import NULL
import encodings
import json
from queue import Full
from venv import create
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import rows_from_range
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Fill, PatternFill, NamedStyle
from fpdf import FPDF
import smtplib
import os
import mimetypes
from email import encoders
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
from email.mime.audio import MIMEAudio
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from dotenv import load_dotenv
import shutil

load_dotenv()

def create_pdf(image_to_pdf, filename, go_to_client):
    
    workBook = load_workbook(filename = filename)
    sheets = workBook.active
    
    full_name = ''
    code = ''
    mail = ''

    for row in range(1, sheets.max_row + 1):
        
        if go_to_client == 1:
            os.mkdir('To_Send')
    
        fist_name = sheets[row][0].value
        if fist_name == None:
            fist_name = ''
        
        name = sheets[row][1].value
        if name == None:
            name = ''
        
        second_name = sheets[row][2].value
        if second_name == None:
            second_name = ''
        
        code = sheets[row][3].value
        if code == None:
            code = ''
        
        mail = sheets[row][4].value
        if mail == None:
            mail = ''
    
        full_name = fist_name + ' ' + name + ' ' + second_name
    
        print(row, full_name, code, mail)

        pdf = FPDF('P', 'mm', 'A4')

        pdf.add_font("Sans", style="", fname="Noto Sans/NotoSans-Regular.ttf", uni=True)
        pdf.add_font("Sans", style="B", fname="Noto Sans/NotoSans-Bold.ttf", uni=True)
        pdf.add_font("Sans", style="I", fname="Noto Sans/NotoSans-Italic.ttf", uni=True)
        pdf.add_font("Sans", style="BI", fname="Noto Sans/NotoSans-BoldItalic.ttf", uni=True)

        pdf.add_page()

        pdf.image(image_to_pdf, x=5, y=5, w=200)
    
        baze_point = 340

        pdf.set_font("Sans", "B", 20)
        pdf.cell(0, baze_point, 'Получил', new_x="LMARGIN", align='C')
        pdf.cell(0, baze_point + 20, full_name, new_x="LMARGIN", align='C')

        pdf.set_font("Sans", "", 16)
        pdf.cell(0, baze_point + 60, 'за участие в XII международном форуме', new_x="LMARGIN", align='C')
        pdf.cell(0, baze_point + 75, '"ОБРАЗОВАНИЕ: РЕАЛИИ И ПЕРСПЕКТИВЫ"', new_x="LMARGIN", align='C')

        pdf.set_font("Sans", "", 14)
        pdf.cell(0, 500, code, new_x="LMARGIN", new_y="NEXT", align='C')
    
        pdf.set_display_mode(zoom='fullpage', layout='continuous')
        
        if go_to_client == 1:
            pdf.output("To_Send/" + full_name + ".pdf")
            
            sender = "msboriskv@gmail.com"
            password = os.getenv("point")
            
            msg = MIMEMultipart()
            msg["From"] = sender
            msg["To"] = mail
            msg["Subject"] = 'Сертификат форума "Образование: Реалии и перспективы"'

            msg.attach(MIMEText("Здравствуйте, " + full_name + " отправляем Вам сертификат форума - Образование: Реалии и перспективы"))

            for file in os.listdir("To_Send"):
                filename = os.path.basename(file)
                ftype, encoding = mimetypes.guess_type(file)
                file_type, subtype = ftype.split("/")
    
                if file_type == "text":
                    with open(f"To_Send/{file}") as f:
                        file = MIMEText(f.read())
                elif file_type == "image":
                    with open(f"To_Send/{file}", "rb") as f:
                        file = MIMEImage(f.read(), subtype)
                elif file_type == "audio":
                    with open(f"To_Send/{file}", "rb") as f:
                        file = MIMEAudio(f.read(), subtype)
                elif file_type == "application":
                    with open(f"To_Send/{file}", "rb") as f:
                        file = MIMEApplication(f.read(), subtype)
                else:
                    with open(f"To_Send"/{file}, "rb") as f:
                        file = MIMEBase(file_type, subtype)
                        file.set_payload(f.read())
                        encoders.encode_base64(file)
    
                file.add_header('Content-Disposition', 'attachment', filename=filename)
                msg.attach(file)

            server = smtplib.SMTP("smtp.gmail.com", 587)
            server.starttls()

            server.login(sender, password)

            server.sendmail(sender, mail, msg.as_string())
            shutil.rmtree('To_Send')
        else:
            pdf.output("Sertificat/" + full_name + ".pdf")

def main():
    go_to_client = int(input("Введите 1 для отпрвки почты или любой символ для сохранения файлов без отправки: "))
    create_pdf("555.png", "Anckett.xlsx", go_to_client = go_to_client)

if __name__ == "__main__":
    main()