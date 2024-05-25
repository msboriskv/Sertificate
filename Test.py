# -*- coding: cp1251 -*-
# -*- coding: utf8 -*-

from asyncio.windows_events import NULL
import encodings
from fileinput import filename
import json
from queue import Full
from venv import create
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.descriptors.base import Float
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import rows_from_range
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Fill, PatternFill, NamedStyle, borders
from fpdf import FPDF
import smtplib
import os
import mimetypes
from email import encoders
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
from email.mime.audio import MIMEAudio
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from dotenv import load_dotenv
import shutil

load_dotenv()

def create_pdf(change):

    if change == 1:
        pass
    
    workBook = load_workbook(filename = 'Для Таблицы/' + 'Таблица.xlsx')
    sheets = workBook.active
    
    full_name = ''
    code = ''
    mail = ''
    
    for row in range(1, sheets.max_row + 1):
        
        pdf = FPDF('P', 'mm', 'A4')
        
        pdf.add_font("Sans", style="", fname="Noto Sans/NotoSans-Regular.ttf", uni=True)
        pdf.add_font("Sans", style="B", fname="Noto Sans/NotoSans-Bold.ttf", uni=True)
        pdf.add_font("Sans", style="I", fname="Noto Sans/NotoSans-Italic.ttf", uni=True)
        pdf.add_font("Sans", style="BI", fname="Noto Sans/NotoSans-BoldItalic.ttf", uni=True)
        
        fist_name = sheets[row][0].value
        if fist_name == None:
            fist_name = ''
        
        name = sheets[row][1].value
        if name == None:
            name = ''
        
        second_name = sheets[row][2].value
        if second_name == None:
            second_name = ''
        
        code = sheets[row][3].value
        if code == None:
            code = ''
        
        mail = sheets[row][4].value
        if mail == None:
            mail = ''
    
        full_name = fist_name + ' ' + name + ' ' + second_name
    
        pdf.set_page_background('555.png')
        pdf.add_page()
        pdf.set_display_mode(zoom='fullpage', layout='continuous')

        pdf.set_font("Sans", "B", 35)
        pdf.y = 148
        pdf.multi_cell(0, 15, 'БЛАГОДАРСТВЕННОЕ\nПИСЬМО', border = 0, align='C')
        pdf.ln()

        pdf.set_font("Sans", "B", 20)
        pdf.y = 185
        pdf.multi_cell(0, 10, 'Уважаемый (ая)\n' + full_name, border = 0, align='C')
        pdf.ln()
        
        pdf.set_font("Sans", "B", 14)
        pdf.y = 205
        pdf.multi_cell(0, 10, code, border = 0, align='C')
        pdf.ln()
    
        pdf.set_font("Sans", "B", 12)
    
        read_txt = open('Text_Thanks.txt', 'r')
        txt_view = read_txt.read()
    
        pdf.y = 220
        pdf.multi_cell(0, 6, txt_view, border = 0, align='C')
        
        pdf.output("На печать/" + full_name + ".pdf")
 

def main():
    change = int(input('Наберите 1 если нужен сертификат и 2 если благодарственное письмо: '))
    create_pdf(change)
    #print('Вы ввели: ' + change)
    
if __name__ == "__main__":
    main()